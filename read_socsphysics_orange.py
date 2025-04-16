import pandas as pd
from openpyxl import load_workbook
import io

def read_excel_socs_cowise(file_path):
    raw_df = pd.read_excel(file_path, header=None)
    header_row = raw_df.iloc[5].copy()
    col_15_value = raw_df.iloc[1, 14]
    header_row.iloc[14] = col_15_value if pd.notna(col_15_value) else "Total"

    df = pd.read_excel(file_path, header=None, skiprows=6, dtype={'SAPID': str})
    df.columns = header_row
    df['SAPID'] = df['SAPID'].astype(str)
    df.columns = df.columns.astype(str).str.strip()
    return df[['SAPID', 'Name of the Student', 'Total']]

def writemarkstoexamtemplate(df, excel_path):
    df['SAPID'] = df['SAPID'].astype(str)
    df = df[~df['SAPID'].astype(str).str.lower().eq('nan')]
    df['SAPID'] = df['SAPID'].astype(float).astype(int)

    df = df.rename(columns={
        'SAPID': 'StudentId',
        'Name of the Student': 'StudentName',
        'Total': 'Marks'
    })

    wb = load_workbook(excel_path)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    id_col = header.index('StudentId') + 1
    name_col = header.index('StudentName') + 1
    marks_col = header.index('Marks') + 1

    for row in range(2, ws.max_row + 1):
        try:
            excel_id = int(ws.cell(row=row, column=id_col).value)
            excel_name = ws.cell(row=row, column=name_col).value
            match = df[df['StudentId'] == excel_id]
            if not match.empty:
                new_marks = match['Marks'].values[0]
                new_marks = 'AB' if new_marks < 0 else str(new_marks)
                ws.cell(row=row, column=marks_col).value = new_marks
                print(f"Updated: {excel_id} ({excel_name}) -> {new_marks}")
        except Exception as e:
            print(f"Error processing row {row}: {e}")

    return wb  # Return workbook for saving in JS context

# Main execution for Pyodide
df = read_excel_socs_cowise("input.xlsx")
wb = writemarkstoexamtemplate(df, "target.xlsx")

# Save to BytesIO for JS to pick up
import io
result_bytes = io.BytesIO()
wb.save(result_bytes)
result_bytes.seek(0)
output = result_bytes.read()

